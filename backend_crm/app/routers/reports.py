import io
from calendar import monthrange
from datetime import date, datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .. import models, schemas
from ..database import get_db
from ..deps import get_current_employee
from ..utils_weeks import get_month_weeks, weekly_max, is_current_week, today_uz
from ..telegram import send_telegram_message
from .attendance import WORK_START_HOUR, WORK_START_MIN, TZ_UZ as TZ_UZ_ATTENDANCE, _note_out

router = APIRouter(prefix="/reports", tags=["Haftalik hisobot"])

_ADMIN_ROLES = {models.RoleEnum.superadmin, models.RoleEnum.direktor, models.RoleEnum.zamdirektor}
_HEAD_ROLES = {models.RoleEnum.bolim_boshligi, models.RoleEnum.boshqarma_boshligi}


def _can_review(reviewer: models.Employee, target: models.Employee) -> bool:
    """reviewer, target xodimning haftalik hisobotini tasdiqlay oladimi."""
    if reviewer.id == target.id:
        return False
    if reviewer.role in _ADMIN_ROLES:
        return target.role in _HEAD_ROLES
    if reviewer.role in _HEAD_ROLES:
        return target.department_id == reviewer.department_id and target.role not in _HEAD_ROLES
    return False


def _reviewer_targets(current: models.Employee, db: Session) -> List[models.Employee]:
    """current tasdiqlay oladigan barcha xodimlar ro'yxati."""
    if current.role in _ADMIN_ROLES:
        return (
            db.query(models.Employee)
            .filter(models.Employee.role.in_(_HEAD_ROLES), models.Employee.is_active == True)
            .order_by(models.Employee.id)
            .all()
        )
    if current.role in _HEAD_ROLES:
        if current.department_id is None:
            return []
        return (
            db.query(models.Employee)
            .filter(
                models.Employee.department_id == current.department_id,
                models.Employee.is_active == True,
                ~models.Employee.role.in_(_HEAD_ROLES),
            )
            .order_by(models.Employee.id)
            .all()
        )
    return []


def _upsert_score(db: Session, emp_id: int, year: int, month: int, creator_id: int) -> models.Score:
    sc = db.query(models.Score).filter(
        models.Score.employee_id == emp_id,
        models.Score.year == year,
        models.Score.month == month,
    ).first()
    if not sc:
        sc = models.Score(employee_id=emp_id, year=year, month=month, created_by=creator_id)
        db.add(sc)
        db.flush()
    return sc


def _sync_bolim_ball(db: Session, emp_id: int, year: int, month: int, actor_id: int):
    """Tasdiqlangan haftalik ballar yig'indisini Score.bolim_ball ga yozadi."""
    total = (
        db.query(models.WeeklyReport)
        .filter(
            models.WeeklyReport.employee_id == emp_id,
            models.WeeklyReport.year == year,
            models.WeeklyReport.month == month,
            models.WeeklyReport.confirmed_at.isnot(None),
        )
        .all()
    )
    s = sum(w.ball or 0 for w in total)
    sc = _upsert_score(db, emp_id, year, month, actor_id)
    sc.bolim_ball = round(s, 2) if total else None


def _week_info_map(year: int, month: int) -> dict[int, dict]:
    return {w["week"]: w for w in get_month_weeks(year, month)}


def _now_uz() -> datetime:
    return datetime.now(TZ_UZ_ATTENDANCE).replace(tzinfo=None)


def _active_override(db: Session, year: int, month: int, week: int) -> Optional[models.WeeklyReportWindowOverride]:
    """Berilgan hafta uchun hozir amalda bo'lgan (muddati o'tmagan) uzaytirish
    yozuvini qaytaradi, bo'lmasa None."""
    row = db.query(models.WeeklyReportWindowOverride).filter(
        models.WeeklyReportWindowOverride.year == year,
        models.WeeklyReportWindowOverride.month == month,
        models.WeeklyReportWindowOverride.week == week,
    ).first()
    if row and row.open_until > _now_uz():
        return row
    return None


def _overrides_map(db: Session, year: int, month: int) -> dict[int, models.WeeklyReportWindowOverride]:
    """Shu oydagi barcha HALI AMALDAGI uzaytirishlarni week -> row shaklida qaytaradi."""
    now = _now_uz()
    rows = db.query(models.WeeklyReportWindowOverride).filter(
        models.WeeklyReportWindowOverride.year == year,
        models.WeeklyReportWindowOverride.month == month,
        models.WeeklyReportWindowOverride.open_until > now,
    ).all()
    return {r.week: r for r in rows}


@router.get("/weeks", response_model=List[schemas.WeekInfo])
def list_weeks(
    year: int,
    month: int,
    _: models.Employee = Depends(get_current_employee),
):
    """Berilgan oydagi haftalar va ularning sana oraliqlari."""
    wmax = weekly_max(year, month)
    return [
        schemas.WeekInfo(
            week=w["week"], start=w["start"], end=w["end"], label=w["label"], max_ball=wmax,
            is_current=is_current_week(w["start"], w["end"]),
        )
        for w in get_month_weeks(year, month)
    ]


MAX_WEEKLY_FILE_BYTES = 10 * 1024 * 1024   # 10MB


def _b64_payload_size_bytes(data_uri: str) -> int:
    """Data-URI'dagi base64 qismining taxminiy dekodlangan hajmi (baytda)."""
    b64 = data_uri.split(",", 1)[-1]
    return (len(b64) * 3) // 4 - b64.count("=")


@router.post("/weekly", response_model=schemas.WeeklyReportOut)
def upload_weekly_report(
    payload: schemas.WeeklyReportUploadIn,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """Joriy foydalanuvchi o'zining haftalik hisobot faylini (va ish tavsifini)
    yuklaydi/yangilaydi — faqat joriy hafta uchun."""
    weeks = _week_info_map(payload.year, payload.month)
    if payload.week not in weeks:
        raise HTTPException(status_code=422, detail="Noto'g'ri hafta raqami")

    w_info = weeks[payload.week]
    override = None
    if not is_current_week(w_info["start"], w_info["end"]):
        override = _active_override(db, payload.year, payload.month, payload.week)
        if not override:
            raise HTTPException(status_code=400, detail="Faqat joriy hafta uchun hisobot yuklash mumkin")

    if payload.file_b64 and _b64_payload_size_bytes(payload.file_b64) > MAX_WEEKLY_FILE_BYTES:
        raise HTTPException(status_code=413, detail="Fayl hajmi 10MB dan oshmasligi kerak")

    rep = db.query(models.WeeklyReport).filter(
        models.WeeklyReport.employee_id == current.id,
        models.WeeklyReport.year == payload.year,
        models.WeeklyReport.month == payload.month,
        models.WeeklyReport.week == payload.week,
    ).first()

    if rep and rep.confirmed_at is not None:
        raise HTTPException(status_code=400, detail="Tasdiqlangan hisobotni qayta yuklab bo'lmaydi")

    if not rep:
        rep = models.WeeklyReport(
            employee_id=current.id, year=payload.year, month=payload.month, week=payload.week,
        )
        db.add(rep)

    rep.description = payload.description
    if payload.file_b64:
        rep.file_name = payload.file_name
        rep.file_b64 = payload.file_b64
    rep.uploaded_at = datetime.utcnow()
    db.commit()
    db.refresh(rep)

    out = schemas.WeeklyReportOut.model_validate(rep)
    out.week_label = w_info["label"]
    out.max_ball = weekly_max(payload.year, payload.month)
    out.is_current = is_current_week(w_info["start"], w_info["end"])
    out.upload_open = True
    out.open_until = override.open_until if override else None
    out.employee_name = current.full_name
    return out


@router.delete("/weekly/{report_id}", status_code=204)
def delete_weekly_report(
    report_id: int,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """Joriy foydalanuvchi o'zining (hali tasdiqlanmagan) haftalik hisobotini
    butunlay o'chiradi — fayl bazadan ham tozalanadi, axlat bo'lib qolmaydi."""
    rep = db.query(models.WeeklyReport).filter(models.WeeklyReport.id == report_id).first()
    if not rep:
        raise HTTPException(status_code=404, detail="Hisobot topilmadi")
    if rep.employee_id != current.id:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")
    if rep.confirmed_at is not None:
        raise HTTPException(status_code=400, detail="Tasdiqlangan hisobotni o'chirib bo'lmaydi")
    db.delete(rep)
    db.commit()


@router.get("/weekly/window-overrides", response_model=List[schemas.WeeklyWindowOverrideOut])
def list_window_overrides(
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """Shu oy uchun berilgan barcha muddat-uzaytirishlar (amal qilish muddati
    o'tganlari ham) — sozlamalar sahifasida ko'rsatish uchun."""
    if current.role not in _ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")
    return (
        db.query(models.WeeklyReportWindowOverride)
        .filter(
            models.WeeklyReportWindowOverride.year == year,
            models.WeeklyReportWindowOverride.month == month,
        )
        .order_by(models.WeeklyReportWindowOverride.week)
        .all()
    )


@router.post("/weekly/window-override", response_model=schemas.WeeklyWindowOverrideOut)
def set_window_override(
    payload: schemas.WeeklyWindowOverrideIn,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """O'tgan (yopilgan) haftaga vaqtincha ochiq muddat beradi — belgilangan
    open_until vaqtigacha o'sha hafta uchun hisobot yuklash yana ruxsat etiladi."""
    if current.role not in _ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")
    weeks = _week_info_map(payload.year, payload.month)
    if payload.week not in weeks:
        raise HTTPException(status_code=422, detail="Noto'g'ri hafta raqami")

    row = db.query(models.WeeklyReportWindowOverride).filter(
        models.WeeklyReportWindowOverride.year == payload.year,
        models.WeeklyReportWindowOverride.month == payload.month,
        models.WeeklyReportWindowOverride.week == payload.week,
    ).first()
    if not row:
        row = models.WeeklyReportWindowOverride(year=payload.year, month=payload.month, week=payload.week)
        db.add(row)
    row.open_until = payload.open_until
    row.created_by = current.id
    db.commit()
    db.refresh(row)
    return row


@router.delete("/weekly/window-override/{override_id}", status_code=204)
def delete_window_override(
    override_id: int,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """Berilgan muddat-uzaytirishni bekor qiladi (haftani muddatidan oldin yopadi)."""
    if current.role not in _ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")
    row = db.query(models.WeeklyReportWindowOverride).filter(models.WeeklyReportWindowOverride.id == override_id).first()
    if row:
        db.delete(row)
        db.commit()


@router.get("/weekly/mine", response_model=List[schemas.WeeklyReportOut])
def my_weekly_reports(
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """Joriy foydalanuvchining shu oydagi haftalik hisobotlari (yuklanmaganlar ham bo'sh hafta sifatida)."""
    weeks = get_month_weeks(year, month)
    wmax = weekly_max(year, month)
    overrides = _overrides_map(db, year, month)
    existing = {
        r.week: r
        for r in db.query(models.WeeklyReport).filter(
            models.WeeklyReport.employee_id == current.id,
            models.WeeklyReport.year == year,
            models.WeeklyReport.month == month,
        ).all()
    }
    result = []
    for w in weeks:
        rep = existing.get(w["week"])
        if rep:
            out = schemas.WeeklyReportOut.model_validate(rep)
        else:
            out = schemas.WeeklyReportOut(
                id=0, employee_id=current.id, year=year, month=month, week=w["week"],
            )
        out.week_label = w["label"]
        out.max_ball = wmax
        out.is_current = is_current_week(w["start"], w["end"])
        override = overrides.get(w["week"])
        out.upload_open = out.is_current or override is not None
        out.open_until = override.open_until if override else None
        out.employee_name = current.full_name
        result.append(out)
    return result


def _build_team_rows(targets: List[models.Employee], year: int, month: int, db: Session) -> List[schemas.WeeklyTeamRowOut]:
    if not targets:
        return []

    weeks = get_month_weeks(year, month)
    wmax = weekly_max(year, month)
    overrides = _overrides_map(db, year, month)
    target_ids = [t.id for t in targets]

    reports = db.query(models.WeeklyReport).filter(
        models.WeeklyReport.employee_id.in_(target_ids),
        models.WeeklyReport.year == year,
        models.WeeklyReport.month == month,
    ).all()
    reports_by_emp: dict[int, dict[int, models.WeeklyReport]] = {}
    for r in reports:
        reports_by_emp.setdefault(r.employee_id, {})[r.week] = r

    scores = {
        sc.employee_id: sc
        for sc in db.query(models.Score).filter(
            models.Score.employee_id.in_(target_ids),
            models.Score.year == year,
            models.Score.month == month,
        ).all()
    }
    dept_map = {d.id: d for d in db.query(models.Department).all()}

    rows = []
    for emp in targets:
        emp_reports = reports_by_emp.get(emp.id, {})
        week_outs = []
        for w in weeks:
            rep = emp_reports.get(w["week"])
            if rep:
                out = schemas.WeeklyReportOut.model_validate(rep)
            else:
                out = schemas.WeeklyReportOut(id=0, employee_id=emp.id, year=year, month=month, week=w["week"])
            out.week_label = w["label"]
            out.max_ball = wmax
            out.is_current = is_current_week(w["start"], w["end"])
            override = overrides.get(w["week"])
            out.upload_open = out.is_current or override is not None
            out.open_until = override.open_until if override else None
            week_outs.append(out)
        dept = dept_map.get(emp.department_id) if emp.department_id else None
        rows.append(schemas.WeeklyTeamRowOut(
            employee_id=emp.id,
            full_name=emp.full_name,
            position=emp.position,
            department_name=dept.name if dept else None,
            weeks=week_outs,
            bolim_ball=scores.get(emp.id).bolim_ball if emp.id in scores else None,
        ))
    return rows


@router.get("/weekly/team", response_model=List[schemas.WeeklyTeamRowOut])
def team_weekly_reports(
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """Joriy foydalanuvchi tasdiqlay oladigan xodimlarning shu oydagi hisobotlari."""
    targets = _reviewer_targets(current, db)
    return _build_team_rows(targets, year, month, db)


@router.get("/weekly/subteam", response_model=List[schemas.WeeklyTeamRowOut])
def subteam_weekly_reports(
    head_id: int,
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """Admin/zamdirektor uchun: berilgan bo'lim boshlig'ining o'z xodimlariga qo'ygan
    haftalik hisobotlari (faqat ko'rish — nazorat/shaffoflik uchun)."""
    if current.role not in _ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")
    head = db.query(models.Employee).filter(models.Employee.id == head_id).first()
    if not head or head.role not in _HEAD_ROLES:
        raise HTTPException(status_code=404, detail="Bo'lim boshlig'i topilmadi")
    targets = _reviewer_targets(head, db)
    return _build_team_rows(targets, year, month, db)


@router.post("/weekly/{report_id}/score", response_model=schemas.WeeklyReportOut)
def score_weekly_report(
    report_id: int,
    payload: schemas.WeeklyReportScoreIn,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """Tasdiqlovchi haftalik hisobotni ko'rib, ball qo'yib tasdiqlaydi."""
    rep = db.query(models.WeeklyReport).filter(models.WeeklyReport.id == report_id).first()
    if not rep:
        raise HTTPException(status_code=404, detail="Hisobot topilmadi")
    if not rep.file_b64:
        raise HTTPException(status_code=400, detail="Hali fayl yuklanmagan")
    if rep.confirmed_at is not None:
        raise HTTPException(status_code=400, detail="Allaqachon tasdiqlangan")

    target = db.query(models.Employee).filter(models.Employee.id == rep.employee_id).first()
    if not target or not _can_review(current, target):
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")

    wmax = weekly_max(rep.year, rep.month)
    if payload.ball < 0 or payload.ball > wmax:
        raise HTTPException(status_code=422, detail=f"Ball 0–{wmax} oralig'ida bo'lishi kerak")

    rep.ball = payload.ball
    rep.confirmed_at = datetime.utcnow()
    rep.confirmed_by = current.id
    db.flush()

    _sync_bolim_ball(db, rep.employee_id, rep.year, rep.month, current.id)
    db.commit()
    db.refresh(rep)

    out = schemas.WeeklyReportOut.model_validate(rep)
    weeks = _week_info_map(rep.year, rep.month)
    w = weeks.get(rep.week)
    if w:
        out.week_label = w["label"]
        out.is_current = is_current_week(w["start"], w["end"])
    out.max_ball = wmax
    out.employee_name = target.full_name
    return out


@router.get("/weekly/file/{report_id}")
def download_weekly_file(
    report_id: int,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """Haftalik hisobot faylini (base64) qaytaradi — egasi yoki tasdiqlovchisi uchun."""
    rep = db.query(models.WeeklyReport).filter(models.WeeklyReport.id == report_id).first()
    if not rep or not rep.file_b64:
        raise HTTPException(status_code=404, detail="Fayl topilmadi")

    if rep.employee_id != current.id:
        target = db.query(models.Employee).filter(models.Employee.id == rep.employee_id).first()
        if not target or (current.role not in _ADMIN_ROLES and not _can_review(current, target)):
            raise HTTPException(status_code=403, detail="Ruxsat yo'q")

    return {"file_name": rep.file_name, "file_b64": rep.file_b64}


def _generate_pending_message(targets: List[models.Employee], db: Session) -> tuple[str, int]:
    """Joriy hafta uchun hisobot yuklamaganlar ro'yxatidan Telegram xabari matnini tayyorlaydi."""
    today = today_uz()
    weeks = get_month_weeks(today.year, today.month)
    current_week = next((w for w in weeks if is_current_week(w["start"], w["end"])), None)
    if not current_week or not targets:
        return "", 0

    target_ids = [t.id for t in targets]
    uploaded_ids = {
        r.employee_id for r in db.query(models.WeeklyReport.employee_id).filter(
            models.WeeklyReport.employee_id.in_(target_ids),
            models.WeeklyReport.year == today.year,
            models.WeeklyReport.month == today.month,
            models.WeeklyReport.week == current_week["week"],
            models.WeeklyReport.file_name.isnot(None),
        ).all()
    }
    pending = [t for t in targets if t.id not in uploaded_ids]
    if not pending:
        return "", 0

    lines = [f"\U0001F4CC {current_week['label']} haftasi uchun hisobot topshirmaganlar:", ""]
    lines += [f"{i}) {p.full_name}" for i, p in enumerate(pending, 1)]
    lines += ["", "Hisobotlarni bugun kechqurun soat 23:59 gacha topshirishingiz so'raladi."]
    return "\n".join(lines), len(pending)


@router.get("/weekly/pending-message", response_model=schemas.PendingMessageOut)
def get_pending_message(
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """Joriy foydalanuvchi nazorat qiladigan xodimlardan joriy hafta hisobot
    topshirmaganlar ro'yxatidan tayyor xabar matnini qaytaradi."""
    targets = _reviewer_targets(current, db)
    text, count = _generate_pending_message(targets, db)
    return schemas.PendingMessageOut(text=text, count=count)


@router.post("/weekly/send-telegram")
def send_weekly_telegram(
    payload: schemas.TelegramMessageIn,
    current: models.Employee = Depends(get_current_employee),
):
    """Telegram guruhiga xabar yuborish (faqat rahbar rollar uchun)."""
    if current.role not in (_ADMIN_ROLES | _HEAD_ROLES):
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")
    if not payload.text.strip():
        raise HTTPException(status_code=422, detail="Xabar matni bo'sh bo'lishi mumkin emas")
    try:
        ok = send_telegram_message(payload.text)
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    if not ok:
        raise HTTPException(status_code=502, detail="Telegram xabarini yuborib bo'lmadi")
    return {"ok": True}


@router.get("/weekly/missing", response_model=schemas.MissingReportOut)
def missing_weekly_reports(
    year: int,
    month: int,
    week: int,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """Berilgan (istalgan — o'tgan yoki joriy) hafta uchun hisobot fayli
    yuklamagan barcha FAOL xodimlar ro'yxati va Telegramga yuborish uchun
    tayyor xabar matni — faqat superadmin/direktor/zamdirektor."""
    if current.role not in _ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")
    weeks = _week_info_map(year, month)
    if week not in weeks:
        raise HTTPException(status_code=422, detail="Noto'g'ri hafta raqami")
    week_label = weeks[week]["label"]

    emps = (
        db.query(models.Employee)
        .filter(models.Employee.is_active == True)
        .order_by(models.Employee.department_id, models.Employee.id)
        .all()
    )
    uploaded_ids = {
        r.employee_id for r in db.query(models.WeeklyReport.employee_id).filter(
            models.WeeklyReport.year == year,
            models.WeeklyReport.month == month,
            models.WeeklyReport.week == week,
            models.WeeklyReport.file_name.isnot(None),
        ).all()
    }
    missing = [e for e in emps if e.id not in uploaded_ids]

    if missing:
        lines = [f"\U0001F4CC {week}-hafta ({week_label}) uchun hisobot topshirmaganlar:", ""]
        lines += [f"{i}) {e.full_name}" for i, e in enumerate(missing, 1)]
        lines += ["", "Iltimos, hisobotingizni tezroq yuklang."]
        text = "\n".join(lines)
    else:
        text = ""

    return schemas.MissingReportOut(
        week=week, week_label=week_label, count=len(missing),
        names=[e.full_name for e in missing], text=text,
    )


# ─── Xodim oylik hisoboti (direktor sahifasi) ─────────────────────────────────

_MONTH_NAMES_UZ = ["Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
                    "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"]

# Haqiqiy ishlagan soat (kelish-ketish oralig'i) qayd etilmaydi — "chiqdim" tugmasi
# hali yo'q. Shu sabab "jami ish soati" standart 8 soatlik kun taxminiga asoslanadi;
# faqat kechikish daqiqalari haqiqiy GPS vaqtidan olinadi.
STANDARD_WORKDAY_MIN = 8 * 60


def _bolim_boshligi_nomi(db: Session, department_id: int | None) -> str | None:
    if department_id is None:
        return None
    head = db.query(models.Employee).filter(
        models.Employee.department_id == department_id,
        models.Employee.role.in_([models.RoleEnum.bolim_boshligi, models.RoleEnum.boshqarma_boshligi]),
    ).first()
    return head.full_name if head else None


@router.get("/monthly/{employee_id}", response_model=schemas.MonthlyReportOut)
def monthly_report(
    employee_id: int,
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """Bitta xodimning bir oylik to'liq hisoboti — davomat, ish vaqti tahlili,
    haftalik hisobotlar va baholash natijalari. Faqat direktor/zamdirektor/superadmin."""
    if current.role not in _ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")

    emp = db.query(models.Employee).filter(models.Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Xodim topilmadi")

    days_in_month = monthrange(year, month)[1]
    today = today_uz()
    last_day_to_count = today.day if (year, month) == (today.year, today.month) else days_in_month

    month_prefix = f"{year:04d}-{month:02d}-"
    attendances = db.query(models.Attendance).filter(
        models.Attendance.employee_id == employee_id,
        models.Attendance.date >= f"{month_prefix}01",
        models.Attendance.date <= f"{month_prefix}{days_in_month:02d}",
    ).all()
    att_by_day = {int(a.date[-2:]): a for a in attendances}

    month_start = f"{month_prefix}01"
    month_end = f"{month_prefix}{days_in_month:02d}"
    notes = db.query(models.AttendanceNote).filter(
        models.AttendanceNote.employee_id == employee_id,
        models.AttendanceNote.date_from <= month_end,
        models.AttendanceNote.date_to >= month_start,
    ).order_by(models.AttendanceNote.date_from).all()
    notes_out = [_note_out(n) for n in notes]

    calendar_days: list[schemas.MonthlyReportCalendarDay] = []
    kelgan = kechikkan = kelmagan = ish_kunlari_jami = 0
    late_minutes_total = 0

    for day in range(1, days_in_month + 1):
        d = date(year, month, day)
        wd = d.weekday()  # 0=Dushanba
        if wd >= 5:
            calendar_days.append(schemas.MonthlyReportCalendarDay(day=day, weekday=wd, status="dam_olish"))
            continue

        ish_kunlari_jami += 1
        if day > last_day_to_count:
            calendar_days.append(schemas.MonthlyReportCalendarDay(day=day, weekday=wd, status="kelajak"))
            continue

        att = att_by_day.get(day)
        if att:
            # check_in DB'da odatda mahalliy (UTC+5) vaqt sifatida saqlanadi (naive) — attendance.py'dagi
            # _to_out bilan bir xil mantiq (tzinfo mavjud bo'lsa ham to'g'ri ishlaydi).
            ci_local = att.check_in.astimezone(TZ_UZ_ATTENDANCE) if att.check_in.tzinfo is not None else att.check_in
            work_start = ci_local.replace(hour=WORK_START_HOUR, minute=WORK_START_MIN, second=0, microsecond=0)
            late = max(0, int(round((ci_local - work_start).total_seconds() / 60.0)))
            status = "kechikkan" if late > 0 else "kelgan"
            if status == "kechikkan":
                kechikkan += 1
                late_minutes_total += late
            else:
                kelgan += 1
            calendar_days.append(schemas.MonthlyReportCalendarDay(day=day, weekday=wd, status=status, time=ci_local.strftime("%H:%M")))
        else:
            kelmagan += 1
            calendar_days.append(schemas.MonthlyReportCalendarDay(day=day, weekday=wd, status="kelmagan"))

    attended_days = kelgan + kechikkan
    jami_ish_soati_min = attended_days * STANDARD_WORKDAY_MIN   # "Jami ish soati" statistika kartochkasi uchun

    # "Ishlash vaqti tahlili" donutining maxraji — shu oyning TALAB QILINADIGAN jami
    # soati (ish_kunlari_jami * 8 soat), attended kunlar emas — shunda halqa oy davomida
    # qancha bajarilganini (yashil) va qolgan qismini (bo'sh) ko'rsatadi.
    donut_total_min = ish_kunlari_jami * STANDARD_WORKDAY_MIN
    kechikish_min = min(late_minutes_total, jami_ish_soati_min)
    samarali_min = max(0, jami_ish_soati_min - kechikish_min)
    pct = lambda part: round(part / donut_total_min * 100, 1) if donut_total_min else 0.0

    # Haftalik hisobotlar
    weeks = get_month_weeks(year, month)
    reports_by_week = {
        r.week: r
        for r in db.query(models.WeeklyReport).filter(
            models.WeeklyReport.employee_id == employee_id,
            models.WeeklyReport.year == year,
            models.WeeklyReport.month == month,
        ).all()
    }
    weekly_rows = []
    for w in weeks:
        r = reports_by_week.get(w["week"])
        weekly_rows.append(schemas.MonthlyReportWeekRow(
            week=w["week"],
            label=f"{w['week']}-hafta ({w['label']})",
            file_name=r.file_name if r else None,
            uploaded_at=r.uploaded_at if r else None,
            ball=r.ball if r else None,
            report_id=r.id if r else None,
        ))

    # Baholash
    sc = db.query(models.Score).filter(
        models.Score.employee_id == employee_id,
        models.Score.year == year,
        models.Score.month == month,
    ).first()
    ijro_ball  = sc.ijro_ball if sc else None
    kadr_ball  = sc.kadr_ball if sc else None
    bolim_ball = sc.bolim_ball if sc else None
    umumiy = None
    if ijro_ball is not None or kadr_ball is not None or bolim_ball is not None:
        umumiy = (ijro_ball or 0) + (kadr_ball or 0) + (bolim_ball or 0)

    dept = emp.department
    hr = db.query(models.Employee).filter(models.Employee.role == models.RoleEnum.kadr).first()

    return schemas.MonthlyReportOut(
        report_id=f"{today.strftime('%y%m%d')}-{employee_id:03d}",
        report_date=today.strftime("%d.%m.%Y"),
        period_label=f"{year}-yil {_MONTH_NAMES_UZ[month - 1]}",
        employee_id=emp.id,
        full_name=emp.full_name,
        position=emp.position,
        department_name=dept.name if dept else None,
        phone=emp.phone,
        has_photo=bool(emp.photo_base64),
        summary=schemas.MonthlyReportSummary(
            kelgan_kunlar=kelgan,
            kechikkan_kunlar=kechikkan,
            kelmagan_kunlar=kelmagan,
            ish_kunlari_jami=ish_kunlari_jami,
            jami_ish_soati_min=jami_ish_soati_min,
        ),
        calendar=calendar_days,
        notes=notes_out,
        weekly_reports=weekly_rows,
        time_analysis=schemas.MonthlyReportTimeAnalysis(
            samarali_min=samarali_min,
            kechikish_min=kechikish_min,
            total_min=donut_total_min,
            samarali_pct=pct(samarali_min),
            kechikish_pct=pct(kechikish_min),
        ),
        scores=schemas.MonthlyReportScores(
            ijro=schemas.MonthlyReportScoreItem(label="Ijro bo'limi bahosi", ball=ijro_ball, max_ball=10),
            kadr=schemas.MonthlyReportScoreItem(label="Kadrlar bo'limi bahosi", ball=kadr_ball, max_ball=25),
            bolim=schemas.MonthlyReportScoreItem(label="Bo'lim boshlig'i bahosi", ball=bolim_ball, max_ball=65),
            umumiy=schemas.MonthlyReportScoreItem(label="Umumiy natija", ball=umumiy, max_ball=100),
            comment=sc.comment if sc else None,
        ),
        prepared_by_name=hr.full_name if hr else None,
        approved_by_name=_bolim_boshligi_nomi(db, emp.department_id),
    )


# ─── Barcha xodimlar oylik davomat jadvali (XLSX ko'rish/yuklab olish) ───────

def _employee_month_days(
    db: Session, employee_id: int, year: int, month: int, days_in_month: int, last_day_to_count: int,
) -> tuple[list[schemas.MonthlyTableCell], int, int, int]:
    """Bitta xodim uchun shu oydagi ISH KUNLARI (dam olish kunlarisiz) bo'yicha
    holat ro'yxati — monthly_report'dagi kalendar mantig'i bilan bir xil, lekin
    faqat ish kunlari qaytariladi va bir nechta xodim uchun jadval qurishga mos."""
    month_prefix = f"{year:04d}-{month:02d}-"
    attendances = db.query(models.Attendance).filter(
        models.Attendance.employee_id == employee_id,
        models.Attendance.date >= f"{month_prefix}01",
        models.Attendance.date <= f"{month_prefix}{days_in_month:02d}",
    ).all()
    att_by_day = {int(a.date[-2:]): a for a in attendances}

    cells: list[schemas.MonthlyTableCell] = []
    kelgan = kechikkan = kelmagan = 0
    for day in range(1, days_in_month + 1):
        d = date(year, month, day)
        if d.weekday() >= 5:
            continue
        if day > last_day_to_count:
            cells.append(schemas.MonthlyTableCell(day=day, status="kelajak"))
            continue
        att = att_by_day.get(day)
        if att:
            ci_local = att.check_in.astimezone(TZ_UZ_ATTENDANCE) if att.check_in.tzinfo is not None else att.check_in
            work_start = ci_local.replace(hour=WORK_START_HOUR, minute=WORK_START_MIN, second=0, microsecond=0)
            late = max(0, int(round((ci_local - work_start).total_seconds() / 60.0)))
            status = "kechikkan" if late > 0 else "kelgan"
            if status == "kechikkan":
                kechikkan += 1
            else:
                kelgan += 1
            cells.append(schemas.MonthlyTableCell(day=day, status=status, time=ci_local.strftime("%H:%M"), late_min=late or None))
        else:
            kelmagan += 1
            cells.append(schemas.MonthlyTableCell(day=day, status="kelmagan"))
    return cells, kelgan, kechikkan, kelmagan


def _build_monthly_table_data(db: Session, year: int, month: int) -> schemas.MonthlyTableOut:
    days_in_month = monthrange(year, month)[1]
    today = today_uz()
    last_day_to_count = today.day if (year, month) == (today.year, today.month) else days_in_month

    working_days = [d for d in range(1, days_in_month + 1) if date(year, month, d).weekday() < 5]
    day_labels = [schemas.MonthlyTableDayLabel(day=d, label=date(year, month, d).strftime("%d.%m")) for d in working_days]

    # "Hamma xodimlar" = barcha FAOL xodimlar — rahbariyat, direktor/zamdirektor ham
    # kiradi (rol bo'yicha filtrlanmaydi, faqat is_active).
    emps = (
        db.query(models.Employee)
        .filter(models.Employee.is_active == True)
        .order_by(models.Employee.department_id, models.Employee.id)
        .all()
    )
    dept_map = {d.id: d.name for d in db.query(models.Department).all()}

    rows: list[schemas.MonthlyTableRow] = []
    for e in emps:
        cells, kelgan, kechikkan, kelmagan = _employee_month_days(db, e.id, year, month, days_in_month, last_day_to_count)
        cell_map = {c.day: c for c in cells}
        rows.append(schemas.MonthlyTableRow(
            employee_id=e.id,
            full_name=e.full_name,
            department_name=dept_map.get(e.department_id),
            cells=[cell_map.get(d) for d in working_days],
            kelgan=kelgan,
            kechikkan=kechikkan,
            kelmagan=kelmagan,
        ))

    return schemas.MonthlyTableOut(
        period_label=f"{year}-yil {_MONTH_NAMES_UZ[month - 1]}",
        days=day_labels,
        rows=rows,
    )


@router.get("/monthly-table", response_model=schemas.MonthlyTableOut)
def monthly_table(
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """Barcha faol xodimlarning shu oydagi kunma-kun davomat jadvali (XLSX
    ko'rish oynasi uchun) — faqat direktor/zamdirektor/superadmin."""
    if current.role not in _ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")
    return _build_monthly_table_data(db, year, month)


def _cell_text(cell: Optional[schemas.MonthlyTableCell]) -> str:
    if cell is None or cell.status == "kelajak":
        return ""
    if cell.status == "kelmagan":
        return "Kelmadi"
    if cell.status == "kechikkan":
        return f"{cell.time} (+{cell.late_min} daq.)"
    return cell.time or ""


@router.get("/monthly-xlsx")
def monthly_xlsx(
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """Barcha faol xodimlarning shu oydagi davomat jadvalini .xlsx fayl sifatida
    yuklab beradi — faqat direktor/zamdirektor/superadmin."""
    if current.role not in _ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")
    table = _build_monthly_table_data(db, year, month)

    wb = Workbook()
    ws = wb.active
    ws.title = table.period_label[:31]

    header = ["Bo'lim", "F.I.Sh."] + [d.label for d in table.days] + ["Kelgan", "Kechikkan", "Kelmagan"]
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    ok_fill     = PatternFill("solid", fgColor="FFE3F7EC")
    late_fill   = PatternFill("solid", fgColor="FFFFF3CD")
    absent_fill = PatternFill("solid", fgColor="FFFDE2E2")

    for row in table.rows:
        ws.append([row.department_name or "", row.full_name] + [_cell_text(c) for c in row.cells] + [row.kelgan, row.kechikkan, row.kelmagan])

    for ri, row in enumerate(table.rows, start=2):
        for ci, cell in enumerate(row.cells, start=3):
            if cell is None:
                continue
            xl_cell = ws.cell(row=ri, column=ci)
            xl_cell.alignment = Alignment(horizontal="center")
            if cell.status == "kelmagan":
                xl_cell.fill = absent_fill
            elif cell.status == "kechikkan":
                xl_cell.fill = late_fill
            elif cell.status == "kelgan":
                xl_cell.fill = ok_fill

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 26
    for i in range(len(table.days) + 3):
        ws.column_dimensions[get_column_letter(3 + i)].width = 14
    ws.freeze_panes = "C2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"oylik_hisobot_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
