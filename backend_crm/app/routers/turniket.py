"""Turniket (badge-reader) xlsx eksportini yuklab, tizimdagi xodimlarga
avtomatik moslashtirib (name_match.py), TurniketAttendance jadvaliga import
qilish — KADR roli uchun. Oqim: /preview (fayl tahlil qilinadi, moslashtirish
natijasi ko'rsatiladi, TurniketImportBatch'ga vaqtincha saqlanadi) -> kadr
mos kelmagan qatorlarni qo'lda tuzatadi -> /commit (shu oy uchun eski
yozuvlar o'chirilib, yangi ma'lumot yoziladi). /tabel — Tizim davomat
(tabel.py/_build_auto_tabel) bilan bir xil ko'rinishdagi oylik jadval, lekin
turniket manbasidan."""
import json
import re
import uuid
from calendar import monthrange
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .. import models, schemas
from ..database import get_db
from ..deps import get_current_employee
from ..name_match import match_employee, normalize
from .attendance import WORK_START_HOUR, WORK_START_MIN

router = APIRouter(prefix="/turniket", tags=["Turniket davomat"])

_IMPORT_ROLES = {models.RoleEnum.kadr, models.RoleEnum.superadmin}
_VIEW_ROLES = {models.RoleEnum.kadr, models.RoleEnum.superadmin,
               models.RoleEnum.direktor, models.RoleEnum.zamdirektor}

_DATE_RE = re.compile(r"(\d{1,2})\.(\d{1,2})\.(\d{4})")
_TIME_RE = re.compile(r"^(\d{1,2}):(\d{2})")


def _cell_time_str(v) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s or s == "-":
            return None
        m = _TIME_RE.match(s)
        return f"{int(m.group(1)):02d}:{m.group(2)}" if m else None
    if hasattr(v, "hour") and hasattr(v, "minute"):
        return f"{v.hour:02d}:{v.minute:02d}"
    return None


def _cell_minutes(v) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s or s == "-":
            return None
        m = _TIME_RE.match(s)
        return int(m.group(1)) * 60 + int(m.group(2)) if m else None
    if hasattr(v, "hour") and hasattr(v, "minute"):
        return v.hour * 60 + v.minute
    return None


def _parse_xlsx(file_bytes: bytes) -> dict:
    """Turniket eksport xlsx: 0-qator sarlavha (har 3 ustunda bitta kun sanasi),
    1-qator "Kirish/Chiqish/Ish vaqti", 2-qatordan boshlab xodimlar."""
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        raise ValueError("Fayl bo'sh yoki noto'g'ri formatda")

    header = rows[0]
    day_cols: List[Tuple[str, int]] = []
    for col_idx in range(2, len(header)):
        v = header[col_idx]
        if v is None:
            continue
        m = _DATE_RE.search(str(v))
        if m:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            day_cols.append((f"{y:04d}-{mo:02d}-{d:02d}", col_idx))

    if not day_cols:
        raise ValueError("Fayldan kunlar aniqlanmadi — sarlavha formatini tekshiring")

    parsed_rows = []
    for r in rows[2:]:
        if len(r) < 2:
            continue
        name = r[1]
        if not name or not str(name).strip():
            continue
        days: Dict[str, dict] = {}
        for date_str, col in day_cols:
            if col + 2 >= len(r):
                continue
            check_in = _cell_time_str(r[col])
            if check_in is None:
                continue
            days[date_str] = {
                "check_in": check_in,
                "check_out": _cell_time_str(r[col + 1]),
                "worked_minutes": _cell_minutes(r[col + 2]),
            }
        parsed_rows.append({"xlsx_name": str(name).strip(), "days": days})

    return {"days_in_month": len(day_cols), "rows": parsed_rows}


@router.post("/preview", response_model=schemas.TurniketPreviewOut)
def preview_import(
    year:    int = Form(...),
    month:   int = Form(...),
    file:    UploadFile = File(...),
    db:      Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    if current.role not in _IMPORT_ROLES:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Faqat .xlsx fayl yuklang")

    raw = file.file.read()
    try:
        parsed = _parse_xlsx(raw)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Faylni o'qib bo'lmadi: {e}")

    employees = db.query(models.Employee).filter(models.Employee.is_active == True).all()
    employees_by_id = {e.id: e for e in employees}
    aliases = {
        a.normalized_name: a.employee_id
        for a in db.query(models.TurniketNameAlias).all()
    }

    preview_rows = []
    for idx, row in enumerate(parsed["rows"]):
        alias_employee_id = aliases.get(normalize(row["xlsx_name"]))
        if alias_employee_id and alias_employee_id in employees_by_id:
            emp, confidence = employees_by_id[alias_employee_id], "saved"
        else:
            emp, confidence = match_employee(row["xlsx_name"], employees)
        preview_rows.append(schemas.TurniketPreviewRow(
            row_index=idx,
            xlsx_name=row["xlsx_name"],
            days_with_data=len(row["days"]),
            matched_employee_id=emp.id if emp else None,
            matched_employee_name=emp.full_name if emp else None,
            confidence=confidence,
        ))
        row["matched_employee_id"] = emp.id if emp else None

    cutoff = datetime.utcnow() - timedelta(hours=2)
    db.query(models.TurniketImportBatch).filter(models.TurniketImportBatch.created_at < cutoff).delete()

    batch_id = uuid.uuid4().hex
    db.add(models.TurniketImportBatch(
        id=batch_id, year=year, month=month,
        payload_json=json.dumps(parsed["rows"]),
        uploaded_by=current.id,
    ))
    db.commit()

    return schemas.TurniketPreviewOut(batch_id=batch_id, days_in_month=parsed["days_in_month"], rows=preview_rows)


@router.post("/commit", response_model=schemas.TurniketCommitOut)
def commit_import(
    data:    schemas.TurniketCommitIn,
    db:      Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    if current.role not in _IMPORT_ROLES:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")

    batch = db.query(models.TurniketImportBatch).filter(models.TurniketImportBatch.id == data.batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Import muddati tugagan, qaytadan yuklang")

    rows = json.loads(batch.payload_json)
    valid_employee_ids = {e.id for e in db.query(models.Employee.id).all()}
    existing_aliases = {a.normalized_name: a for a in db.query(models.TurniketNameAlias).all()}

    resolved: Dict[int, dict] = {}
    skipped = 0
    for idx, row in enumerate(rows):
        employee_id = data.corrections.get(idx, row.get("matched_employee_id"))
        if employee_id is None or employee_id not in valid_employee_ids:
            skipped += 1
            continue
        resolved.setdefault(employee_id, {}).update(row["days"])

        norm_name = normalize(row["xlsx_name"])
        if norm_name:
            existing = existing_aliases.get(norm_name)
            if existing:
                existing.employee_id = employee_id
            else:
                alias = models.TurniketNameAlias(normalized_name=norm_name, employee_id=employee_id)
                db.add(alias)
                existing_aliases[norm_name] = alias

    if not resolved:
        db.delete(batch)
        db.commit()
        return schemas.TurniketCommitOut(imported_employees=0, imported_days=0, skipped_rows=skipped)

    month_prefix = f"{batch.year:04d}-{batch.month:02d}-"
    db.query(models.TurniketAttendance).filter(
        models.TurniketAttendance.employee_id.in_(resolved.keys()),
        models.TurniketAttendance.date.like(f"{month_prefix}%"),
    ).delete(synchronize_session=False)

    imported_days = 0
    for employee_id, days in resolved.items():
        for date_str, cell in days.items():
            db.add(models.TurniketAttendance(
                employee_id=employee_id, date=date_str,
                check_in=cell.get("check_in"), check_out=cell.get("check_out"),
                worked_minutes=cell.get("worked_minutes"),
            ))
            imported_days += 1

    db.delete(batch)
    db.commit()

    return schemas.TurniketCommitOut(
        imported_employees=len(resolved), imported_days=imported_days, skipped_rows=skipped,
    )


_TABEL_EXCLUDED_ROLES = {
    models.RoleEnum.superadmin, models.RoleEnum.direktor, models.RoleEnum.zamdirektor,
}
_TABEL_EXCLUDED_STATUSES = {
    models.EmployeeStatusEnum.shafyor_farrosh, models.EmployeeStatusEnum.dekret,
}
_STATUS_RANGE_CODE = {
    models.EmployeeStatusEnum.otpuska:            "MT",
    models.EmployeeStatusEnum.oquv_tatilida:       "O'",
    models.EmployeeStatusEnum.xizmat_safarida:     "K",
    models.EmployeeStatusEnum.mehnatga_layoqatsiz: "B",
}


def _build_turniket_tabel(db: Session, year: int, month: int) -> schemas.AutoTabelOut:
    """Tizim davomat (tabel.py/_build_auto_tabel) bilan bir xil ko'rinishdagi
    oylik jadval, lekin manba — turniket import (TurniketAttendance). Kelgan
    kun katagida "8" o'rniga haqiqiy kirish vaqti (masalan "8:35") ko'rsatiladi."""
    days_in_month = monthrange(year, month)[1]

    depts = db.query(models.Department).all()
    dept_order = {d.id: d.order_num for d in depts}
    dept_map = {d.id: d.name for d in depts}

    employees = (
        db.query(models.Employee)
        .filter(
            models.Employee.role.notin_(_TABEL_EXCLUDED_ROLES),
            models.Employee.status.notin_(_TABEL_EXCLUDED_STATUSES),
        )
        .all()
    )
    employees.sort(key=lambda e: (dept_order.get(e.department_id, 9999), e.full_name))
    emp_ids = [e.id for e in employees]

    month_prefix = f"{year:04d}-{month:02d}-"
    records = db.query(models.TurniketAttendance).filter(
        models.TurniketAttendance.employee_id.in_(emp_ids),
        models.TurniketAttendance.date >= f"{month_prefix}01",
        models.TurniketAttendance.date <= f"{month_prefix}{days_in_month:02d}",
    ).all()
    by_emp_day: Dict[int, Dict[int, models.TurniketAttendance]] = {}
    for r in records:
        by_emp_day.setdefault(r.employee_id, {})[int(r.date[-2:])] = r

    rows = []
    for emp in employees:
        emp_days = by_emp_day.get(emp.id, {})
        cells: Dict[str, str] = {}
        worked_min = 0
        late_min = 0
        for day in range(1, days_in_month + 1):
            d = date(year, month, day)
            if d.weekday() >= 5:
                cells[str(day)] = "X"
                continue

            in_status_range = bool(
                emp.status_date_from and emp.status_date_to
                and emp.status_date_from <= d.isoformat() <= emp.status_date_to
            )
            rec = emp_days.get(day)
            if in_status_range and emp.status in _STATUS_RANGE_CODE:
                cells[str(day)] = _STATUS_RANGE_CODE[emp.status]
            elif rec is not None and rec.check_in:
                cells[str(day)] = "8"
                worked_min += rec.worked_minutes or 0
                h, m = map(int, rec.check_in.split(":"))
                late_min += max(0, (h * 60 + m) - (WORK_START_HOUR * 60 + WORK_START_MIN))
            else:
                cells[str(day)] = ""

        rows.append(schemas.AutoTabelRow(
            employee_id=emp.id,
            full_name=emp.full_name,
            department_id=emp.department_id,
            department_name=dept_map.get(emp.department_id),
            cells=cells,
            worked_min=worked_min,
            late_min=late_min,
        ))

    working_days = sum(1 for d in range(1, days_in_month + 1) if date(year, month, d).weekday() < 5)
    return schemas.AutoTabelOut(days_in_month=days_in_month, working_days=working_days, rows=rows)


@router.get("/tabel", response_model=schemas.AutoTabelOut)
def get_turniket_tabel(
    year:    int,
    month:   int,
    db:      Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    if current.role not in _VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")
    return _build_turniket_tabel(db, year, month)


@router.get("/employee/{employee_id}", response_model=schemas.TurniketEmployeeMonthOut)
def get_employee_month(
    employee_id: int,
    year:        int,
    month:       int,
    db:          Session = Depends(get_db),
    current:     models.Employee = Depends(get_current_employee),
):
    """Bitta xodimning bir oylik turniket kalendari — har kun uchun kirish VA
    chiqish vaqti (Turniket davomat jadvalida xodim ustiga bosilganda ochiladi)."""
    if current.role not in _VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")

    emp = db.query(models.Employee).filter(models.Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Xodim topilmadi")
    dept = db.query(models.Department).filter(models.Department.id == emp.department_id).first() if emp.department_id else None

    days_in_month = monthrange(year, month)[1]
    month_prefix = f"{year:04d}-{month:02d}-"
    records = db.query(models.TurniketAttendance).filter(
        models.TurniketAttendance.employee_id == employee_id,
        models.TurniketAttendance.date >= f"{month_prefix}01",
        models.TurniketAttendance.date <= f"{month_prefix}{days_in_month:02d}",
    ).all()
    rec_by_day = {int(r.date[-2:]): r for r in records}

    days: List[schemas.TurniketDayDetail] = []
    kelgan_kunlar = 0
    kelmagan_kunlar = 0
    jami_ish_soati_min = 0
    for day in range(1, days_in_month + 1):
        d = date(year, month, day)
        weekday = d.weekday()
        rec = rec_by_day.get(day)

        in_status_range = bool(
            emp.status_date_from and emp.status_date_to
            and emp.status_date_from <= d.isoformat() <= emp.status_date_to
        )
        if weekday >= 5:
            status = "dam_olish"
        elif in_status_range and emp.status in _STATUS_RANGE_CODE:
            status = "status_" + _STATUS_RANGE_CODE[emp.status]
        elif rec is not None and rec.check_in:
            status = "kelgan"
            kelgan_kunlar += 1
            jami_ish_soati_min += rec.worked_minutes or 0
        elif d <= date.today():
            status = "kelmagan"
            kelmagan_kunlar += 1
        else:
            status = "kelajak"

        days.append(schemas.TurniketDayDetail(
            day=day, weekday=weekday, status=status,
            check_in=rec.check_in if rec else None,
            check_out=rec.check_out if rec else None,
            worked_minutes=rec.worked_minutes if rec else None,
        ))

    return schemas.TurniketEmployeeMonthOut(
        employee_id=emp.id,
        full_name=emp.full_name,
        position=emp.position,
        department_name=dept.name if dept else None,
        days=days,
        kelgan_kunlar=kelgan_kunlar,
        kelmagan_kunlar=kelmagan_kunlar,
        jami_ish_soati_min=jami_ish_soati_min,
    )
