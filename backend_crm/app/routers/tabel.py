import io
from calendar import monthrange
from datetime import date, datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .. import models, schemas
from ..database import get_db
from ..deps import get_current_employee
from .attendance import TZ_UZ, WORK_START_HOUR, WORK_START_MIN

router = APIRouter(prefix="/tabel", tags=["Tabel"])


def _is_admin(emp: models.Employee) -> bool:
    return emp.role in {models.RoleEnum.superadmin, models.RoleEnum.direktor, models.RoleEnum.zamdirektor}


@router.get("/month", response_model=List[schemas.TabelMonthRecord])
def get_month_tabel(
    year: int,
    month: int,
    department_id: int | None = None,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """
    Bir oylik tabelni qaytaradi.
    - superadmin/direktor/zamdirektor: istalgan bo'lim yoki hamma
    - bolim_boshligi/boshqarma_boshligi: faqat o'z bo'limi
    """
    if _is_admin(current):
        dept_id = department_id  # admin may specify any dept
    else:
        dept_id = current.department_id  # restricted to own dept

    # Xodimlar ro'yxati
    q = db.query(models.Employee).filter(models.Employee.is_active == True)
    if dept_id is not None:
        q = q.filter(models.Employee.department_id == dept_id)
    employees = q.order_by(models.Employee.id).all()

    # Mavjud yozuvlar
    emp_ids = [e.id for e in employees]
    records = (
        db.query(models.TabelRecord)
        .filter(
            models.TabelRecord.employee_id.in_(emp_ids),
            models.TabelRecord.year == year,
            models.TabelRecord.month == month,
        )
        .all()
    )
    # {emp_id: {day: code}}
    rec_map: dict[int, dict[int, str]] = {}
    for r in records:
        rec_map.setdefault(r.employee_id, {})[r.day] = r.code

    result = []
    for emp in employees:
        result.append(schemas.TabelMonthRecord(
            employee_id=emp.id,
            full_name=emp.full_name,
            position=emp.position,
            work_rate=emp.work_rate,
            days=rec_map.get(emp.id, {}),
        ))
    return result


@router.post("/save", response_model=dict)
def save_tabel(
    payload: schemas.TabelBatchIn,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """Ko'p kunni bir vaqtda saqlash (upsert)."""
    for rec in payload.records:
        existing = (
            db.query(models.TabelRecord)
            .filter(
                models.TabelRecord.employee_id == rec.employee_id,
                models.TabelRecord.year == rec.year,
                models.TabelRecord.month == rec.month,
                models.TabelRecord.day == rec.day,
            )
            .first()
        )
        if existing:
            existing.code = rec.code
            existing.created_by = current.id
        else:
            db.add(models.TabelRecord(
                employee_id=rec.employee_id,
                year=rec.year,
                month=rec.month,
                day=rec.day,
                code=rec.code,
                created_by=current.id,
            ))
    db.commit()
    return {"saved": len(payload.records)}


_STATUS_RANGE_CODE = {
    models.EmployeeStatusEnum.otpuska:             "MT",
    models.EmployeeStatusEnum.oquv_tatilida:        "O'",
    models.EmployeeStatusEnum.xizmat_safarida:      "K",
    models.EmployeeStatusEnum.mehnatga_layoqatsiz:  "B",
}


_AUTO_TABEL_EXCLUDED_ROLES = {
    models.RoleEnum.superadmin, models.RoleEnum.direktor, models.RoleEnum.zamdirektor,
}
_AUTO_TABEL_EXCLUDED_STATUSES = {
    models.EmployeeStatusEnum.shafyor_farrosh, models.EmployeeStatusEnum.dekret,
}
STANDARD_WORKDAY_MIN = 8 * 60


def _fmt_hm(total_min: int) -> str:
    h, m = divmod(max(0, total_min), 60)
    return f"{h} soat {m} daqiqa" if m else f"{h} soat"


def _build_auto_tabel(db: Session, year: int, month: int) -> schemas.AutoTabelOut:
    days_in_month = monthrange(year, month)[1]
    today = datetime.now(TZ_UZ).date()
    last_day_to_count = today.day if (year, month) == (today.year, today.month) else days_in_month
    working_days = sum(1 for d in range(1, days_in_month + 1) if date(year, month, d).weekday() < 5)

    depts = db.query(models.Department).all()
    dept_order = {d.id: d.order_num for d in depts}
    dept_map = {d.id: d.name for d in depts}

    employees = (
        db.query(models.Employee)
        .filter(
            models.Employee.role.notin_(_AUTO_TABEL_EXCLUDED_ROLES),
            models.Employee.status.notin_(_AUTO_TABEL_EXCLUDED_STATUSES),
        )
        .all()
    )
    employees.sort(key=lambda e: (dept_order.get(e.department_id, 9999), e.full_name))
    emp_ids = [e.id for e in employees]

    month_prefix = f"{year:04d}-{month:02d}-"
    attendances = db.query(models.Attendance).filter(
        models.Attendance.employee_id.in_(emp_ids),
        models.Attendance.date >= f"{month_prefix}01",
        models.Attendance.date <= f"{month_prefix}{days_in_month:02d}",
    ).all()
    att_by_emp_day: dict[int, dict] = {}
    for a in attendances:
        att_by_emp_day.setdefault(a.employee_id, {})[int(a.date[-2:])] = a

    rows = []
    for emp in employees:
        emp_days = att_by_emp_day.get(emp.id, {})
        cells: dict[str, str] = {}
        worked_min = 0
        late_min = 0
        for day in range(1, days_in_month + 1):
            d = date(year, month, day)
            if d.weekday() >= 5:
                cells[str(day)] = "X"
                continue
            if day > last_day_to_count:
                cells[str(day)] = ""
                continue

            in_status_range = bool(
                emp.status_date_from and emp.status_date_to
                and emp.status_date_from <= d.isoformat() <= emp.status_date_to
            )
            att = emp_days.get(day)
            if in_status_range and emp.status in _STATUS_RANGE_CODE:
                cells[str(day)] = _STATUS_RANGE_CODE[emp.status]
            elif att is not None:
                cells[str(day)] = "8"
                worked_min += STANDARD_WORKDAY_MIN
                ci_local = att.check_in.astimezone(TZ_UZ) if att.check_in.tzinfo is not None else att.check_in
                work_start = ci_local.replace(hour=WORK_START_HOUR, minute=WORK_START_MIN, second=0, microsecond=0)
                late_min += max(0, int(round((ci_local - work_start).total_seconds() / 60.0)))
            else:
                cells[str(day)] = ""

        rows.append(schemas.AutoTabelRow(
            employee_id=emp.id,
            full_name=emp.full_name,
            department_id=emp.department_id,
            department_name=dept_map.get(emp.department_id),
            cells=cells,
            worked_min=worked_min,
            late_min=late_min,
        ))

    return schemas.AutoTabelOut(days_in_month=days_in_month, working_days=working_days, rows=rows)


_AUTO_TABEL_VIEW_ROLES = {
    models.RoleEnum.kadr, models.RoleEnum.superadmin,
    models.RoleEnum.direktor, models.RoleEnum.zamdirektor,
}


@router.get("/auto", response_model=schemas.AutoTabelOut)
def get_auto_tabel(
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """Attendance (GPS) va Employee.status asosida avtomatik hisoblangan oylik
    davomat jadvali — kadr/superadmin/direktor/zamdirektor uchun. Direktor,
    zamdirektor va texnik xodimlar (shafyor/farrosh) ro'yxatga kirmaydi.
    Kod: "8" (kelgan), "X" (dam olish kuni), "MT" (mehnat ta'tili), "O'" (o'quv
    ta'tili), "K" (xizmat safari), "B" (bolnichniy), "Д" (dekret), "" (bo'sh)."""
    if current.role not in _AUTO_TABEL_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")
    return _build_auto_tabel(db, year, month)


@router.get("/auto-xlsx")
def auto_tabel_xlsx(
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """Xodimlar davomati jadvalini .xlsx fayl sifatida yuklab beradi."""
    if current.role not in _AUTO_TABEL_VIEW_ROLES:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")
    data = _build_auto_tabel(db, year, month)
    required_min = data.working_days * STANDARD_WORKDAY_MIN

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month:02d}.{year}"[:31]

    header = ["Ism familiyasi"] + [str(d) for d in range(1, data.days_in_month + 1)] + ["Jami ish soati", "Kechikkan vaqti"]
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    fills = {
        "8":  PatternFill("solid", fgColor="FFE3F7EC"),
        "X":  PatternFill("solid", fgColor="FFF4F9FD"),
        "MT": PatternFill("solid", fgColor="FFFFF3CD"),
        "O'": PatternFill("solid", fgColor="FFEDE9FB"),
        "K":  PatternFill("solid", fgColor="FFE3EEFF"),
        "B":  PatternFill("solid", fgColor="FFFDE2E2"),
        "Д":  PatternFill("solid", fgColor="FFF0F0F0"),
    }

    for row in data.rows:
        ws.append(
            [row.full_name] + [row.cells.get(str(d), "") for d in range(1, data.days_in_month + 1)]
            + [f"{_fmt_hm(row.worked_min)}/{_fmt_hm(required_min)}", _fmt_hm(row.late_min)]
        )

    for ri, row in enumerate(data.rows, start=2):
        for d in range(1, data.days_in_month + 1):
            code = row.cells.get(str(d), "")
            if not code:
                continue
            cell = ws.cell(row=ri, column=1 + d)
            cell.alignment = Alignment(horizontal="center")
            fill = fills.get(code)
            if fill:
                cell.fill = fill
        for extra_col in (2 + data.days_in_month, 3 + data.days_in_month):
            ws.cell(row=ri, column=extra_col).alignment = Alignment(horizontal="center")
            ws.cell(row=ri, column=extra_col).font = Font(bold=True)

    ws.column_dimensions["A"].width = 26
    for i in range(data.days_in_month):
        ws.column_dimensions[get_column_letter(2 + i)].width = 6
    ws.column_dimensions[get_column_letter(2 + data.days_in_month)].width = 16
    ws.column_dimensions[get_column_letter(3 + data.days_in_month)].width = 16
    ws.freeze_panes = "B2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"xodimlar_davomati_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.delete("/clear")
def clear_month(
    year: int,
    month: int,
    department_id: int,
    db: Session = Depends(get_db),
    current: models.Employee = Depends(get_current_employee),
):
    """Bo'lim xodimlari uchun bir oyni tozalash."""
    emp_ids = [
        e.id for e in db.query(models.Employee.id)
        .filter(models.Employee.department_id == department_id)
        .all()
    ]
    deleted = (
        db.query(models.TabelRecord)
        .filter(
            models.TabelRecord.employee_id.in_(emp_ids),
            models.TabelRecord.year == year,
            models.TabelRecord.month == month,
        )
        .delete(synchronize_session=False)
    )
    db.commit()
    return {"deleted": deleted}
